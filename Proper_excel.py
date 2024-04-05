 import pandas as pd
from openpyxl import load_workbook

# Define the chunk size
chunksize = 10000  # Adjust this value based on your available memory

# Initialize an empty list to store the processed chunks
chunks = []

# Open the Excel file using openpyxl
wb = load_workbook("your_excel_file.xlsx", read_only=True)
ws = wb.active

# Read the data in chunks
for row in ws.iter_rows(min_row=2, values_only=True, max_row=ws.max_row, max_col=ws.max_column):
    chunk.append(row)
    if len(chunk) >= chunksize:
        # Process the chunk
        chunk_df = pd.DataFrame(chunk, columns=ws[1])
        # Your processing code here (e.g., pivot, merge, fillna)
        # Append the processed chunk to the list
        chunks.append(chunk_df)
        # Clear the chunk for the next iteration
        chunk = []

# Process the remaining data in the last chunk
if chunk:
    chunk_df = pd.DataFrame(chunk, columns=ws[1])
    # Your processing code here (e.g., pivot, merge, fillna)
    # Append the processed chunk to the list
    chunks.append(chunk_df)

# Concatenate the processed chunks into a single DataFrame
final_df = pd.concat(chunks)

# Save the final DataFrame as an Excel file
final_df.to_excel("output_file.xlsx", index=False)
